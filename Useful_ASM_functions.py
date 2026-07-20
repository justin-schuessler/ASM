import sys
import json
import datetime
import requests
from requests.auth import HTTPBasicAuth
import subprocess
import get_1password_field
from openpyxl import Workbook

Base_URL = "https://platform.sedaraapis.com/"
asm_out_path = "C:/Users/justin.schuessler/Documents/API_Output/Sedara_ASM/"


def list_asm_integrations(headers, tenant_id, params=None):
    if params is None:
        params = {"page": 1, "limit": 100, "pretty": "true", "include_health_report": "true", "include_configuration_report": "true", "include_disabled": "true"}

    response = requests.get(f"{Base_URL}" + "collector/v1/" + tenant_id + "/integrations", headers=headers, params=params)
    print(response.status_code)
    #print(response.headers)
    # print(response.text)
    response.raise_for_status()
    return response.json()


def get_active_integrations(headers, tenant_id, params=None):
    all_tenant_integrations = list_asm_integrations(headers=headers, tenant_id=tenant_id, params=params)
    tenant_integration_dict_list = all_tenant_integrations.get("data")
    configured_integrations_list = {}
    not_configured_integration_list = {}

    for integration in tenant_integration_dict_list:
        integration_out_dict = {}
        enabled = integration["attributes"].get("enabled")
        name = integration["attributes"].get("title")
        configured = integration["meta"].get("configured")
        description = integration["attributes"].get("description")
        integration_id = integration.get("id")
        integration_out_dict["name"] = name
        integration_out_dict["description"] = description
        if configured > 0:
            config_report = integration["meta"].get("configuration_reports")
            recent_conf_report = config_report[0]
            conf_report_attributes = recent_conf_report.get("attributes")

            health_report = integration["meta"].get("health_report")
            health_rep_attrib = health_report.get("attributes")

            status_list = conf_report_attributes.get("status")
            state = health_rep_attrib.get("state")
            last_seen = integration["meta"].get("last_seen_at")
            healthy = False

            integration_out_dict["status_list"] = status_list
            integration_out_dict["state"] = state
            integration_out_dict["last_seen"] = last_seen

            for status in status_list:
                if status == "Healthy":
                    healthy = True

            if state and healthy:
                healthy = True

            integration_out_dict["healthy"] = healthy
            configured_integrations_list[integration_id] = integration_out_dict

        if configured == 0:
            not_configured_integration_list[integration_id] = integration_out_dict

    tenant_configuration_stats_dict = {}
    tenant_configuration_stats_dict["configured_integrations_list"] = configured_integrations_list
    tenant_configuration_stats_dict["not_configured_integration_list"] = not_configured_integration_list
    return tenant_configuration_stats_dict


# forbiden? receive "you probably should not be here" message
def list_asm_organizations(headers, params=None):
    if params is None:
        params = {"page": 1, "per_page": 100}

    response = requests.get(f"{Base_URL}" + "dataservices/v2/organizations", headers=headers, params=params)
    print(response.status_code)
    print(response.headers)
    print(response.text)
    response.raise_for_status()

    return response.json()


def list_asm_tenants(headers, params=None):
    if params is None:
        params = {"page": 1, "size": 300}

    response = requests.get(f"{Base_URL}" + "tenant-management/tenants", headers=headers, params=params)
    #print(response.status_code)
    #print(response.headers)
    # print(response.text)
    response.raise_for_status()

    return response.json()


def display_tenants(headers, params=None):
    tenants_return = list_asm_tenants(headers, params)
    tenants_list = tenants_return.get("tenants")
    for t in tenants_list:
        print(f"Name:{t['attributes'].get('name')}  |Hidden: {t['attributes'].get('hidden')}  |Type: {t['attributes'].get('type')}  |Kind: {t.get('kind')}  |ID: {t.get('id')} ")
    return tenants_list


def get_active_tenants(headers, params=None):
    tenants_return = list_asm_tenants(headers, params)
    tenants_list = tenants_return.get("tenants")
    active_tenants = {}
    for t in tenants_list:
        if not (t['attributes'].get('hidden')) and t['attributes'].get('type') == "CUSTOMER" and t.get('kind') == "platform_TENANT":
            active_tenants[t.get('id')] = t['attributes'].get('name')

    return active_tenants


# Return Empty List Currently
def list_all_asm_collectors_health(headers, params=None):
    if params is None:
        params = {"pretty": "true"}

    response = requests.get(f"{Base_URL}" + "collector/v1/collectors/health", headers=headers, params=params)
    print(response.status_code)
    print(response.headers)
    # print(response.text)
    response.raise_for_status()

    return response.json()


def out_to_excel(data, xl_save_path):
    wb = Workbook()
    ws = wb.active
    row = 2

    ws["A1"] = "Tenant ID"
    ws["B1"] = "Tenant Name"
    ws["C1"] = "Integration Name"
    ws["D1"] = "Integration ID"
    ws["E1"] = "Healthy"
    ws["F1"] = "State"
    ws["G1"] = "Status List"
    ws["H1"] = "Last Seen"
    data_keys = data.keys()

    for key in data_keys:
        configed_dict = data[key].get("configured_integrations_list")
        if len(configed_dict) > 0:
            for configed_integration in data[key].get("configured_integrations_list").keys():
                ws.cell(row=row, column=1).value = key
                ws.cell(row=row, column=2).value = data[key].get("name")
                ws.cell(row=row, column=3).value = data[key]["configured_integrations_list"][configed_integration].get("name")
                ws.cell(row=row, column=4).value = configed_integration
                ws.cell(row=row, column=5).value = data[key]["configured_integrations_list"][configed_integration].get("healthy")
                ws.cell(row=row, column=6).value = data[key]["configured_integrations_list"][configed_integration].get("state")
                status_string = ", ".join(data[key]["configured_integrations_list"][configed_integration].get("status_list"))
                ws.cell(row=row, column=7).value = status_string
                ws.cell(row=row, column=8).value = data[key]["configured_integrations_list"][configed_integration].get("last_seen")
                row += 1

    wb.save(xl_save_path)


def get_base_url():
    return Base_URL


def get_asm_out_path():
    return asm_out_path


def get_asm_platform_api_access_key():
    return get_1password_field.get_1password_field("Sedara Platform ASM API", "Access key", "API")
